import time
import requests
import pandas as pd
import io
import openpyxl
import GrammarBasedParserR

# Replace with your actual URL and headers
BASE_URL ="https://ttl.demo.configit.cloud:8080/api/v1/wi/"
API_KEY = 'YTAwMjNlMzkwZmY4NGJhN2I2YWUzMjZhOTBlMTM5NzdfZGVlYzJkY2UyNjA3NDI2Y2IyZDk3MGM2OTllNmNlNTE='


# Headers for the API requests;
HEADERS = {'Authorization': f'Apikey {API_KEY}','Content-Type': 'application/json','Accept': 'application/json'}

push_families = True
push_features = True
push_rules = False
push_views = False
push_products = True
push_rules = False

def create_work_item():
    work_item_data = {
        "name": "Automotive Config",
        "description": "Automotive Config",
        "assignedUsernames":["sundaram.shanmuga@tatatechnologies.com"]
    }
    response = requests.post(BASE_URL, headers=HEADERS, json=work_item_data)
    print(f"Response status code: {response.status_code}")
    print(f"Response text: {response.text}")
    
    if response.status_code == 200:
        response_data = response.json()
        return response_data.get("id")
    else:
        print(f"Failed to create work item. Status code: {response.status_code}")
        return None


def send_requests(data_structure, wi_id):
    try:
        product_code = "LUXURY VEHICLE"
        product_description = "LUXURY VEHICLE"
        content = data_structure.content

        #create families
        for family in content:
            payload = {
                "familyType": "Feature",
                "code": family.code,
                "description": family.description,
                "lifecycle": "Enabled",
                "labels": ["PLM"]
            }
            if push_families:
                response = requests.post(f"{BASE_URL}{wi_id}/library/families", headers=HEADERS, json=payload)
                print(f"{response.status_code}Added Family {family.code} to library \tStatus Code: {response.status_code}, Response: {response.json()}")   

            #create features
            for feature in family.features:
                payload = {
                    "familyType": "Feature",
                    "code": feature.code,
                    "familyCode" : family.code,
                    "description": feature.description,
                    "lifecycle": "Enabled"
                }
                if push_features:
                    response = requests.post(f"{BASE_URL}{wi_id}/library/features", headers=HEADERS, json=payload)
                    print(f"{response.status_code}\tAdded Feature {feature.code} to library \tStatus Code: {response.status_code}, Response: {response.json()}")   

        # create product
        BRAND = "FF"
        payload = {
            "code": product_code,
            "description": product_description,
            "featureCode": product_code,
            "brandCode": BRAND.replace(" ", "_").upper(),
            "useArithmeticRules": True,
            "useBrochureModels": False,
            "useAdvancedEffectivity": False,
            "canBeUsedAsSubmodel": False
        }
        if push_products:
            response = requests.post(f"{BASE_URL}{wi_id}/products/productmodels", headers=HEADERS, json=payload)
            print(f"{response.status_code}\tAdded product model {product_code} \tStatus Code: {response.status_code}, Response: {response.json()}")

        # add families
        for family in content:
            description = family.description
            code = family.code
            payload = {
                "familyType": "Feature",
                "code": code,
                "isCalculated": False,
                "isPrivate": False,
                "effectivities": [
                    {
                        "startEventCode": f"{product_code}_START",
                        "endEventCode": f"{product_code}_END"
                    }
                ]
            }
            if push_families:
                response = requests.post(f"{BASE_URL}{wi_id}/products/productmodels/{product_code}/families", headers=HEADERS, json=payload)
                print(f"{response.status_code}\tAdded {family.code} to product {product_code} for \tStatus Code: {response.status_code}, Response: {response.json()}")

        # create rules
        if push_rules:
            for rule in data_structure.rules:
                response = requests.post(f"{BASE_URL}{wi_id}/products/productmodels/{product_code}/rules", headers=HEADERS, json=generate_rule_payload(rule, product_code))
                print(f"{response.status_code}\tAdded rule {rule.rule_id} to product {product_code}\tStatus Code: {response.status_code}, Response: {response.json()}")

        if push_views:
            response = requests.post(f"{BASE_URL}{wi_id}/products/productmodels/{product_code}/views", headers=HEADERS, json=generate_view_payload(data_structure))
            print(f"{response.status_code}\tAdded view to product {product_code}\tStatus Code: {response.status_code}, Response: {response.json()}")

    except Exception as e:
        print(f"An error occurred: {e}")

def generate_rule_payload(rule, product_code):
    payload = []
    rule_payload = {
        "text": rule.expression,
        "description": rule.description,
        "type": "Text", 
        "code": rule.rule_id,
        "intent": "Engineering",  
        "isEnabled": True,
        "isLocked": False,
        "effectivity": 
                {
                    "startEventCode": f"{product_code}_START",
                    "endEventCode": f"{product_code}_END"
                },
        "labels": ["PLM"],
        "isPropagated": False
    }
    return rule_payload

def generate_view_payload(data_structure):
    payload = {
        "code": "STANDARD",
        "description": "Standard View",
        "sections": []
    }

    for section in data_structure.view:
        section_payload = {
            "code": section.description.replace(" ", "_").upper(),
            "description": section.description.replace("Group ", ""),
            "type": "Standard",
            "sections": [],
            "contentCodes": [family.code for family in section.families]
        }
        payload["sections"].append(section_payload)

    return payload

class Feature:
    def __init__(self, code, description):
        self.code = code
        self.description = description

class Family:
    def __init__(self, code, description):
        self.code = code
        self.description = description
        self.features = []

class Section:
    def __init__(self, description):
        self.description = description
        self.families = []

class Rule:
    def __init__(self, rule_type, rule_id, description, expression):
        self.rule_type = rule_type
        self.rule_id = rule_id
        self.description = description
        self.expression = expression

class DataStructure:
    def __init__(self):
        self.content = []
        self.view = []
        self.rules = []

def create_id_mapping(data_structure):
    """
    Creates an ID mapping array for the given data structure.

    :param data_structure: The DataStructure object containing families, features, and rules.
    :return: A dictionary mapping descriptions to their unique IDs.
    """
    id_mapping = {}

    # Map families and their features using their codes as IDs
    for family in data_structure.content:
        id_mapping[family.description] = family.code
        for feature in family.features:
            id_mapping[feature.description] = feature.code
    return id_mapping


def fetch_family(wi_id: str, product_model: str, family_code: str):
    """Fetch a family from Configit API and return JSON or None."""
    url = f"{BASE_URL}{wi_id}/products/productmodels/{product_model}/families/{family_code}"
    resp = requests.get(url, headers=HEADERS)
    if resp.status_code == 200:
        return resp.json()
    else:
        print(f"Failed to fetch family {family_code}: {resp.status_code} {resp.text}")
        return None


def fetch_features_for_family(wi_id: str, product_model: str, family_code: str):
    """Fetch features for a family and return list of feature JSON objects."""
    url = f"{BASE_URL}{wi_id}/products/productmodels/{product_model}/families/{family_code}/features"
    resp = requests.get(url, headers=HEADERS)
    if resp.status_code == 200:
        return resp.json()
    else:
        print(f"Failed to fetch features for family {family_code}: {resp.status_code} {resp.text}")
        return []


def build_data_structure_from_configit(wi_id: str, product_model: str, family_codes: list):
    """Given work item id, product model, and list of family codes, build DataStructure."""
    ds = DataStructure()
    for code in family_codes:
        fam_json = fetch_family(wi_id, product_model, code)
        if not fam_json:
            continue
        fam = Family(code=fam_json.get("code"), description=fam_json.get("description"))
        # fetch features
        feats = fetch_features_for_family(wi_id, product_model, code)
        # API may return list or object with 'features' key
        if isinstance(feats, dict) and "features" in feats:
            feat_list = feats.get("features", [])
        else:
            feat_list = feats

        for f in feat_list:
            fcode = f.get("code")
            fdesc = f.get("description")
            fam.features.append(Feature(code=fcode, description=fdesc))

        ds.content.append(fam)

    # create a default view/section
    section = Section(description="Configit Extracted")
    section.families = ds.content.copy()
    ds.view.append(section)
    return ds

def read_rules(id_mapping, file_path: str):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rules = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rule_type = row[2]  # Column C
        rule_id = row[3]    # Column D
        if row[8]:
            description =  row[8] # Column I
        else:
            description = rule_id

        expression = row[7].replace("\"\"", "\"")  # Column H
        
        parsed_rule = GrammarBasedParserR.transform_rules([expression], id_mapping)
        for rule in parsed_rule:
            print("Parsed rule\t" + expression + " \t to \t " + rule)
            if rule_type and rule_id and expression:
                rules.append(Rule(rule_type=rule_type, rule_id=rule_id, description=description, expression=rule))

    return rules

def read_content_new_format(file_path: str):
    """
    Reads content from an Excel file with a different format where:
    - Family Description column contains option names (may repeat)
    - Family Code column contains option codes
    - System Description column contains choice descriptions
    - Feature Code column contains choice codes
    """
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Initialize the data structure
    data_structure = DataStructure()
    
    # Find column indices
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    family_desc_idx = header_row.index("Family Description") if "Family Description" in header_row else None
    family_code_idx = header_row.index("Family Code") if "Family Code" in header_row else None
    feature_desc_idx = header_row.index("System Description") if "System Description" in header_row else None
    feature_code_idx = header_row.index("Feature Code") if "Feature Code" in header_row else None
    
    if None in (family_desc_idx, family_code_idx, feature_desc_idx, feature_code_idx):
        raise ValueError("Required columns not found in the Excel file")
    
    # Dictionary to keep track of families by code
    families_dict = {}
    
    # Process data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        family_desc = row[family_desc_idx]
        family_code = row[family_code_idx]
        feature_desc = row[feature_desc_idx]
        feature_code = row[feature_code_idx]
        
        # Skip if any required field is empty
        if not all([family_desc, family_code, feature_desc, feature_code]):
            continue
        
        # Create or get family
        if family_code not in families_dict:
            family = Family(code=family_code, description=family_desc)
            families_dict[family_code] = family
            data_structure.content.append(family)
        else:
            family = families_dict[family_code]
        
        # Add feature to family
        feature = Feature(code=feature_code, description=feature_desc)
        family.features.append(feature)
    
    # Create a default section containing all families
    # You can modify this if you have section information in your Excel
    default_section = Section(description="Default Group")
    default_section.families = list(families_dict.values())
    data_structure.view.append(default_section)
    
    return data_structure

def read_content(file_path: str):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Initialize the data structure
    data_structure = DataStructure()

    current_section = None
    current_family = None

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        cell_value = row[1].strip() if row[1] else None  # Column B
        if cell_value is None:
            continue

        # Detect sections (Group names)
        if cell_value.startswith("Group"):
            if current_section:
                data_structure.view.append(current_section)
            current_section = Section(description=cell_value.strip())

        # Detect families (Design Options)
        elif cell_value.startswith("Design Option"):
            family_code, description = parse_code_and_description(cell_value)
            current_family = Family(code="O"+family_code, description=description)
            data_structure.content.append(current_family)
            if current_section:
                current_section.families.append(current_family)

        # Detect features (Design Choices)
        elif cell_value.startswith("Design Choice"):
            if current_family:
                feature_code, description = parse_code_and_description(cell_value)
                feature = Feature(code="C"+feature_code, description=description)
                current_family.features.append(feature)
        
        else:
            print("unknown cellvalue:" + cell_value)

    # Append the last section
    if current_section:
        data_structure.view.append(current_section)

    return data_structure

def parse_code_and_description(text):
    # Split the text into code and description
    parts = text.split(",", 1)
    code = parts[0].split()[-1].strip()
    description = parts[1].strip() if len(parts) > 1 else ""
    return code, description

def print_statistics(data_structure):
    num_features = sum(len(family.features) for family in data_structure.content)
    num_families = len(data_structure.content)
    num_sections = len(data_structure.view)
    num_rules = len(data_structure.rules)

    print(f"Statistics:")
    print(f"  Total Features: {num_features}")
    print(f"  Total Families: {num_families}")
    print(f"  Total Sections: {num_sections}")
    print(f"  Total Rules: {num_rules}")
def main():
    print("Start Migration")
    content_file_path = "Book 7.xlsx"
    rules_file_path = "Book 8.xlsx"

    result = read_content_new_format(content_file_path)
    id_mapping = create_id_mapping(result)
    #result.rules = read_rules(id_mapping, rules_file_path)
    
    print_statistics(result)
    print("Content read")
    wi_id = create_work_item()
    print("Work item created:" + str(wi_id))
    send_requests(result, wi_id)
    print("Migration finished")

if __name__ == "__main__":
    main()